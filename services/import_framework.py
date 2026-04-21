from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime


def normalize_header_key(header: str | None) -> str:
    if header is None:
        return ""
    s = str(header).strip().lower()
    s = re.sub(r"[\s\-]+", "_", s)
    s = re.sub(r"[^\w_]+", "", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _cell_to_scalar(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    return val


def parse_csv_bytes(raw: bytes) -> list[list]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [list(row) for row in reader]


def parse_xlsx_bytes(raw: bytes) -> list[list]:
    from openpyxl import load_workbook

    bio = io.BytesIO(raw)
    wb = load_workbook(bio, read_only=True, data_only=True)
    try:
        ws = wb.active
        out: list[list] = []
        for row in ws.iter_rows(values_only=True):
            out.append([_cell_to_scalar(c) for c in row])
        return out
    finally:
        wb.close()


def load_tabular_upload(filename: str, raw: bytes) -> list[list]:
    fn = (filename or "").lower()
    if fn.endswith(".csv"):
        return parse_csv_bytes(raw)
    if fn.endswith((".xlsx", ".xlsm")):
        try:
            return parse_xlsx_bytes(raw)
        except Exception as exc:  # pragma: no cover - openpyxl details vary
            raise ValueError(f"Could not read Excel file: {exc}") from exc
    raise ValueError("Use a .csv, .xlsx, or .xlsm file (export from Excel or Google Sheets).")


def detect_header_row(
    grid: list[list],
    *,
    header_mapper,
    min_matches: int = 2,
    scan_limit: int = 50,
) -> tuple[int | None, list[dict[str, object]]]:
    for ri, row in enumerate(grid[:scan_limit]):
        mapped: list[dict[str, object]] = []
        for ci, cell in enumerate(row):
            if cell is None or str(cell).strip() == "":
                continue
            raw_header = str(cell).strip()
            normalized = normalize_header_key(raw_header)
            suggested = header_mapper(raw_header)
            if suggested:
                mapped.append(
                    {
                        "index": ci,
                        "header": raw_header,
                        "normalized": normalized,
                        "suggested": suggested,
                    }
                )
        if len(mapped) >= min_matches:
            return ri, mapped
    return None, []


def rows_from_mapping(
    data_rows: list[list],
    mapping: dict[int, str],
    *,
    min_required_field: str | None = None,
    header_row_index: int = 0,
) -> list[dict[str, str]]:
    out_rows: list[dict[str, str]] = []
    for offset, row in enumerate(data_rows, start=1):
        if not row or all((c is None or str(c).strip() == "") for c in row):
            continue
        d: dict[str, str] = {}
        for ci, field_name in mapping.items():
            if not field_name:
                continue
            value = row[ci] if ci < len(row) else ""
            d[field_name] = "" if value is None else str(value).strip()
        if min_required_field and not (d.get(min_required_field) or "").strip():
            continue
        d["_sheet_row"] = str(header_row_index + offset + 1)
        out_rows.append(d)
    return out_rows
