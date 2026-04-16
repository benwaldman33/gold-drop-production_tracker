"""Spreadsheet → Purchase import: header aliases and file parsing (CSV / Excel)."""
from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime

# Normalized header (lowercase, underscores) → canonical field name on Purchase (+ strain for lot)
PURCHASE_IMPORT_HEADER_ALIASES: dict[str, str] = {}
_aliases_groups: dict[str, list[str]] = {
    "supplier": [
        "supplier", "farm", "source", "vendor", "grower", "supplier_name", "producer",
        "seller", "farm_name",
    ],
    "purchase_date": [
        "purchase_date", "date", "order_date", "po_date", "purchase", "buy_date",
    ],
    "paid_date": [
        "paid_date", "pay_date", "date_paid", "payment_date",
    ],
    "purchase_week": [
        "week", "fiscal_week", "report_week", "budget_week", "purchase_week",
    ],
    "payment_method": [
        "payment_method", "pay_method", "payment_type",
    ],
    "total_cost": [
        "total_cost", "amount", "total_amount", "invoice_amount", "paid_amount",
        "cost", "payment_total",
    ],
    "stated_weight_lbs": [
        "stated_weight_lbs", "weight_lbs", "lbs", "stated_lbs", "quantity_lbs", "weight",
        "est_weight_lbs", "biomass_lbs", "total_lbs", "pounds", "qty_lbs",
        "invoice_weight",
    ],
    "actual_weight_lbs": [
        "actual_weight_lbs", "actual_lbs", "received_lbs", "received_weight_lbs",
        "net_weight_lbs", "actual_weight",
    ],
    "batch_id": [
        "batch_id", "batch", "batchid", "manifest", "manifest_id", "lot_id", "tag",
    ],
    "delivery_date": [
        "delivery_date", "expected_delivery", "ship_date", "eta", "delivery",
        "expected_delivery_date",
    ],
    "status": [
        "status", "purchase_status", "order_status",
    ],
    "stated_potency_pct": [
        "stated_potency_pct", "potency", "thc_pct", "stated_potency", "pct_potency",
        "estimated_potency", "est_potency", "potency_pct", "thca_pct",
    ],
    "tested_potency_pct": [
        "tested_potency_pct", "tested_potency", "lab_potency", "coa_potency",
    ],
    "price_per_lb": [
        "price_per_lb", "price_lb", "cost_per_lb", "pp_lb", "per_lb",
    ],
    "harvest_date": [
        "harvest_date", "harvest",
    ],
    "storage_note": [
        "storage_note", "storage", "warehouse_note",
    ],
    "license_info": [
        "license_info", "license", "licence", "license_number",
    ],
    "coa_status_text": [
        "coa_status_text", "coa_status", "coa", "testing_status_text",
    ],
    "notes": [
        "notes", "note", "comment", "comments", "description",
    ],
    "queue_placement": [
        "queue_placement", "queue", "placement",
    ],
    "clean_or_dirty": [
        "clean_or_dirty", "clean_dirty",
    ],
    "indoor_outdoor": [
        "indoor_outdoor", "indoor_outdoor", "grow_type", "cultivation",
    ],
    "strain": [
        "strain", "strain_name", "variety", "cultivar", "strain_s",
    ],
}
for _canon, _labels in _aliases_groups.items():
    for _lbl in _labels:
        PURCHASE_IMPORT_HEADER_ALIASES[_lbl] = _canon


def normalize_header_key(header: str | None) -> str:
    if header is None:
        return ""
    s = str(header).strip().lower()
    s = re.sub(r"[\s\-]+", "_", s)
    s = re.sub(r"[^\w_]+", "", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def header_to_canonical(header: str | None) -> str | None:
    """Map a single column header to canonical field name, or None to skip column."""
    key = normalize_header_key(header)
    if not key:
        return None
    if key in PURCHASE_IMPORT_HEADER_ALIASES:
        return PURCHASE_IMPORT_HEADER_ALIASES[key]
    # Direct match to canonical name
    if key in _aliases_groups:
        return key
    return None


def _cell_to_scalar(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    return val


def parse_csv_bytes(raw: bytes) -> list[list]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [list(row) for row in reader]


def parse_xlsx_bytes(raw: bytes) -> list[list]:
    from openpyxl import load_workbook

    bio = io.BytesIO(raw)
    wb = load_workbook(bio, read_only=True, data_only=True)
    try:
        ws = wb.active
        out: list[list] = []
        for row in ws.iter_rows(values_only=True):
            out.append([_cell_to_scalar(c) for c in row])
        return out
    finally:
        wb.close()


def parse_purchase_spreadsheet_upload(filename: str, raw: bytes) -> tuple[list[dict[str, str]], list[str]]:
    """
    Returns (rows as dict canonical_field -> string value, warnings).
    First non-empty row = headers; following rows = data. Empty rows skipped.
    """
    warnings: list[str] = []
    fn = (filename or "").lower()
    if fn.endswith(".csv"):
        grid = parse_csv_bytes(raw)
    elif fn.endswith((".xlsx", ".xlsm")):
        try:
            grid = parse_xlsx_bytes(raw)
        except Exception as e:
            raise ValueError(f"Could not read Excel file: {e}") from e
    else:
        raise ValueError("Use a .csv, .xlsx, or .xlsm file (export from Excel or Google Sheets).")

    if not grid:
        raise ValueError("The file appears to be empty.")

    # Find header row (first row with at least one mapped column)
    header_idx = None
    header_map: list[tuple[int, str]] = []  # col index -> canonical

    for ri, row in enumerate(grid[:50]):
        mapped: list[tuple[int, str]] = []
        for ci, cell in enumerate(row):
            if cell is None or str(cell).strip() == "":
                continue
            canon = header_to_canonical(str(cell).strip())
            if canon:
                mapped.append((ci, canon))
        if len(mapped) >= 2:  # need supplier + something meaningful
            header_idx = ri
            header_map = mapped
            break

    if header_idx is None or not header_map:
        raise ValueError(
            "Could not detect a header row with recognizable columns. "
            "Include at least columns for supplier (or Farm/Source) and purchase date or weight."
        )

    canon_names = {c for _, c in header_map}
    if "supplier" not in canon_names:
        raise ValueError("A supplier column is required (header such as Supplier, Farm, or Source).")
    if "purchase_date" not in canon_names and "stated_weight_lbs" not in canon_names:
        warnings.append("No Purchase date or Weight column detected; rows may fail validation.")

    out_rows: list[dict[str, str]] = []
    for ri in range(header_idx + 1, len(grid)):
        row = grid[ri]
        if not row or all((c is None or str(c).strip() == "") for c in row):
            continue
        d: dict[str, str] = {}
        for ci, canon in header_map:
            if ci < len(row):
                v = row[ci]
                if v is None:
                    d[canon] = ""
                else:
                    d[canon] = str(v).strip()
            else:
                d[canon] = ""
        # Skip completely blank supplier rows
        if not (d.get("supplier") or "").strip():
            continue
        d["_sheet_row"] = str(ri + 1)  # 1-based for display
        out_rows.append(d)

    if not out_rows:
        raise ValueError("No data rows found below the header.")

    if len(out_rows) > 2000:
        raise ValueError("Maximum 2000 data rows per upload; split the file and try again.")

    return out_rows, warnings
